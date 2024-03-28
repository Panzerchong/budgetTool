from gettext import translation
import math
from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse,JsonResponse,QueryDict
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .models import Margins, Product_Price, RateTable, Project,BoM, RateTableCost,Service,BillOfMaterials,ServiceCategory,BOMCategory, Vendor, Summary,Profile
from .forms import BomCategoryModelForm, OrderForm, ProductPriceModelForm, ProjectForm, ProjectModelForm,BillModelForm, RateCostModelForm, ServiceCategoryModelForm,ServiceModelForm, VendorModelForm,UserForm
import json
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.formula import *
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from.decorators import allowed_users, unauthenticated_user
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.db.models import Case, When, Value, IntegerField

percentage_style = NamedStyle(name='percentage')
percentage_style.number_format = '0.00%'

currency_style = NamedStyle(name='currency')
currency_style.number_format = '$#,##0'

border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'), 
    bottom=Side(border_style='thin', color='000000') 
)
currency_style.border = border
percentage_style.border = border

@receiver(post_save, sender=User)
def create_user_profile(sender, instance, created, **kwargs):
    if created:
        Profile.objects.create(user=instance)

@receiver(post_save, sender=User)
def save_user_profile(sender, instance, **kwargs):
    instance.profile.save()


def registerPage(request):
	if request.method == 'POST':
		form = UserForm(request.POST)
		if form.is_valid():
			form.save()
			username = form.cleaned_data['username']
			password = form.cleaned_data['password1']
			user = authenticate(username=username, password=password)
			login(request, user)
			messages.success(request, f'{username}, You Have Successfully Registered!')
			return redirect("/budgetTool")
	else:
		form = UserForm()
		return render(request, 'budgetTool/user/register.html', {'form':form})
	return render(request, 'budgetTool/user/register.html', {'form':form})

def change_password(request):
    user=User.objects.get(username=request.user)
    print(user)
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Your password was successfully updated!')
            return redirect("/budgetTool")
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'budgetTool/user/change_password.html', {'form': form})

@unauthenticated_user
def login_user(request):
    if request.method == 'POST':
        username=request.POST['username']
        password=request.POST['password']

        user=authenticate(request, username=username, password=password)
        if user is not None:
            login(request,user)
            return redirect("/budgetTool")
        else:
            messages.success(request,"There was an error log in. Please try again...")
            return render(request, 'budgetTool/budget_list.html',{})

def logout_user(request):
    logout(request)
    messages.success(request,"You have been logged out...")
    return redirect("/budgetTool")

def budget_list(request):
    templates = Project.objects.filter(isTemplate=True)
    projects=Project.objects.filter(isTemplate=False)
    context={
        "projects": projects,
        "templates": templates,
    }
    return render(request, 'budgetTool/budget_list.html',context)

@allowed_users(allowed_roles=['Admin','Account Manager','Finance','Engineering Lead'])
def rate_table(request):
    user_profile=Profile.objects.get(user=request.user)
    table_order=user_profile.RateTable_index
    if table_order == None:
        table_order = ""
    table_ids = [int(i) for i in table_order if i.isdigit()]
    table = RateTableCost.objects.all()
    preserved_order = Case(
        *[When(pk=pk, then=Value(i)) for i, pk in enumerate(table_ids)],
        default=Value(len(table_ids)+1),
        output_field=IntegerField()
    )
    tableCost = table.annotate(custom_order=preserved_order).order_by('custom_order')

    context={
        "table": table,
        "tableCost": tableCost,
    }
    return render(request,'budgetTool/rate_table.html',context)

def budget_detail(request,pk):
    user_profile = Profile.objects.get(user=request.user)

    bom_order = user_profile.bom_index
    if bom_order == None:
        bom_order = ""
    bom_ids = [int(i) for i in bom_order if i.isdigit()]
    bom_items = BillOfMaterials.objects.filter(project_id=pk)
    preserved_order = Case(
        *[When(pk=pk, then=Value(i)) for i, pk in enumerate(bom_ids)],
        default=Value(len(bom_ids)+1),
        output_field=IntegerField()
    )
    bom = bom_items.annotate(custom_order=preserved_order).order_by('custom_order')

    service_order = user_profile.service_index
    if service_order == None:
        service_order = ""
    service_ids = [int(i) for i in service_order if i.isdigit()]
    service_items = Service.objects.filter(project_id=pk)
    service_preserved_order = Case(
        *[When(pk=pk, then=Value(i)) for i, pk in enumerate(service_ids)],
        default=Value(len(service_ids)+1),
        output_field=IntegerField()
    )
    service = service_items.annotate(custom_order=service_preserved_order).order_by('custom_order')

    project=Project.objects.get(id=pk)
    table=RateTable.objects.all()
    # bom=BillOfMaterials.objects.filter(project_id=pk)
    #filter category that is used by current project bom
    bom_category=BOMCategory.objects.filter(project_BOM_category__in=bom).distinct()
    #service
    # service=Service.objects.filter(project_id=pk)
    #filter out category not used
    service_category=ServiceCategory.objects.filter(project_service_category__in=service).distinct()

    bomForm=BillModelForm()
    if request.method =="POST":
        print("received")
        bomForm=BillModelForm(request.POST)
        if bomForm.is_valid():
            bomForm.save()
            print("created a new bom item")
            return JsonResponse({'status': 'success'})
        else:
            print("invalid form")
            errors = bomForm.errors.as_json()
            return JsonResponse({'status': 'error', 'errors': errors}, status=400)
        
    if len(bom)+len(service) == 0:
        list_margin=100
        list_adjusted_margin=100
        actual_margin=100
        quoted_margin=100
    else:
        list_margin=(1-(project.cost_est_bom+project.cost_est_service)/(project.list_bom+project.list_service))*100
        list_adjusted_margin=(1-(project.cost_adjusted_bom+project.cost_adjusted_service)/(project.list_adjusted_bom+project.list_adjusted_service))*100
        actual_margin=(1-(project.actual_bom+project.actual_service)/(project.quote_BOM+project.quote_Service))*100
        quoted_margin=(1-(project.cost_est_bom+project.cost_est_service)/(project.quote_BOM+project.quote_Service))*100

    summary = Summary(
        cost_est=project.cost_est_bom+project.cost_est_service,
        cost_adjusted=project.cost_adjusted_bom+project.cost_adjusted_service,
        list=project.list_bom+project.list_service,
        list_adjusted=project.list_adjusted_bom+project.list_adjusted_service,
        actual=project.actual_bom+project.actual_service,
        list_margin=list_margin,
        list_adjusted_margin=list_adjusted_margin,
        actual_margin=actual_margin,
        quoted_margin=quoted_margin,
    )

    if len(bom)==0:
        margin_est_bom=100
        margin_adjusted_bom=100
        margin_actual_bom=100
        margin_quoted_bom=100
    else:
        margin_est_bom=round((1-(project.cost_est_bom/project.list_bom))*100,2)
        margin_adjusted_bom=round((1-(project.cost_adjusted_bom/project.list_adjusted_bom))*100,2)
        margin_actual_bom=round((1-(project.actual_bom/project.quote_BOM))*100,2)
        margin_quoted_bom=round((1-(project.cost_est_bom/project.quote_BOM))*100,2)
    
    if len(service)==0:
        margin_est_service=100
        margin_adjusted_service=100
        margin_actual_service=100
    else:
        margin_est_service=round((1-(project.cost_est_service/project.list_service))*100,2)
        margin_adjusted_service=round((1-(project.cost_adjusted_service/project.list_adjusted_service))*100,2)
        margin_actual_service=round((1-(project.actual_service/project.quote_Service))*100,2)

    print(margin_est_service)
    margin=Margins(
        margin_est_bom=margin_est_bom,
        margin_est_service=margin_est_service,
        margin_adjusted_bom=margin_adjusted_bom,
        margin_adjusted_service=margin_adjusted_service,
        margin_actual_bom=margin_actual_bom,
        margin_actual_service=margin_actual_service,
        margin_quoted_bom=margin_quoted_bom,
        # margin_est_bom=round((1-(project.cost_est_bom/project.list_bom))*100,2),
        # margin_est_service=round((1-(project.cost_est_service/project.list_service))*100,2),
        # margin_adjusted_bom=round((1-(project.cost_adjusted_bom/project.list_adjusted_bom))*100,2),
        # margin_adjusted_service=round((1-(project.cost_adjusted_service/project.list_adjusted_service))*100,2),
        # margin_actual_bom=round((1-(project.actual_bom/project.quote_BOM))*100,2),
        # margin_actual_service=round((1-(project.actual_service/project.quote_Service))*100,2),
        # margin_quoted_bom=round((1-(project.cost_est_bom/project.quote_BOM))*100,2),
    )

    context={
        "summary":summary,
        "margin":margin,
        "table":table,
        "project":project,
        "service": service,
        "bom":bom,
        "service_category": service_category,
        "existing_bom_category":bom_category,
        "bomForm":BillModelForm(),
    }
    return render(request, 'budgetTool/budget_detail.html',context)

##create new template
def create_project(request):
    form=ProjectModelForm()
    if request.method =="POST":
        form=ProjectModelForm(request.POST)
        if form.is_valid():
            project=form.save()
            print("created a new template")
            return redirect(f'/budgetTool/{project.pk}')
    context={
        "form":ProjectModelForm()
    }
    return render(request,"budgetTool/create_project.html",context)

def editProject(request,pk):
    project=Project.objects.get(id=pk)

    if request.method == 'POST':
        if request.POST.get('name'):
            project.name=request.POST.get('name','')
        if request.POST.get('quote_BOM'):
            project.quote_BOM=request.POST.get('quote_BOM','')
        if request.POST.get('quote_Service'):
            project.quote_Service=request.POST.get('quote_Service','')
        if request.POST.get('adjust_Service'):
            project.adjust_Service=request.POST.get('adjust_Service','')
            for service in Service.objects.filter(project=project):
                service.hours_adjusted = service.hours_estimated * (1 + float(project.adjust_Service))
                service.sub_total_adjusted_list = service.hours_adjusted * service.rate_list + service.travel_estimate
                service.sub_total_adjusted_cost_est = service.hours_adjusted * service.rate_cost + service.travel_estimate
                service.save()
        if request.POST.get('adjust_BOM'):
            project.adjust_BOM=request.POST.get('adjust_BOM','')
        if request.POST.get('travel_weekly'):
            project.travel_weekly=request.POST.get('travel_weekly','')
            for service in Service.objects.filter(project=project):
                if "On Site" in service.type:
                    service.travel_estimate = service.hours_estimated * float(project.travel_weekly) / 40
                else:
                    service.travel_estimate = 0

                service.sub_total_list = service.hours_estimated * service.rate_list + service.travel_estimate
                service.sub_total_adjusted_list = service.hours_adjusted * service.rate_list + service.travel_estimate
                service.sub_total_cost_est = service.hours_estimated * service.rate_cost + service.travel_estimate
                service.sub_total_adjusted_cost_est = service.hours_adjusted * service.rate_cost + service.travel_estimate
                service.save()
        project.save()
        return HttpResponse("Project updated successfully")
    
    ProjectForm=ProjectModelForm(instance=project)
    context = {"ProjectForm": ProjectForm, "project": project}
    if request.GET.get('quote_BOM'):
        return render(request,'budgetTool/partials/project/editProjectBomQuote.html',context)
    elif request.GET.get('name'):
        return render(request,'budgetTool/partials/project/editProjectName.html',context)
    elif request.GET.get('quote_Service'):
        return render(request,'budgetTool/partials/project/editProjectServiceQuote.html',context)
    elif request.GET.get('adjust_Service'):
        return render(request,'budgetTool/partials/project/editProjectServiceAdjust.html',context)
    elif request.GET.get('adjust_BOM'):
        return render(request,'budgetTool/partials/project/editProjectBomAdjust.html',context)
    elif request.GET.get('travel_weekly'):
        return render(request,'budgetTool/partials/project/editProjectTravel.html',context)
    else:
        return HttpResponse("Project update")
   
##calculate the summary after bom and service is updated
def project_update(pk):
    project=Project.objects.get(id=pk)
    boms=BillOfMaterials.objects.filter(project=project)
    services=Service.objects.filter(project=project)
    # print(f'check boms: {boms}')
    ##clear old data
    project.cost_est_bom=0
    project.cost_adjusted_bom=0
    project.list_bom=0
    project.list_adjusted_bom=0
    project.actual_bom=0
    project.cost_est_service=0
    project.cost_adjusted_service=0
    project.list_service=0
    project.list_adjusted_service=0
    project.actual_service=0
    project.hours=0
    project.hours_adjusted=0
    
    for bom in boms:
        ## cost estimate
        project.cost_est_bom+=bom.estimate_cost*bom.quantity
        ## list
        project.list_bom+=bom.sales_price*bom.quantity
        ## actual
        project.actual_bom+=bom.actual_cost*bom.quantity
    ## cost adjusted
    project.cost_adjusted_bom=project.cost_est_bom
    ## list adjusted
    project.list_adjusted_bom=project.list_bom*(1+project.adjust_BOM)

    for service in services:
        ## cost estimate
        project.cost_est_service+=service.sub_total_cost_est
        ## cost adjusted
        project.cost_adjusted_service+=service.sub_total_adjusted_cost_est
        ## list
        project.list_service+=service.sub_total_list
        ## list adjusted
        project.list_adjusted_service+=service.sub_total_adjusted_list
        ## actual
        project.actual_service+=service.cost_actual
        ## hours
        project.hours+=service.hours_estimated
        ## hours adjusted
        project.hours_adjusted+=service.hours_adjusted
    print(f'hours adjusted: {project.hours_adjusted}')
    project.save()

def project_delete(request,pk):
    project=Project.objects.get(id=pk)
    project.delete()
    return redirect("/budgetTool")

def create_bom(request, pk):
    project = Project.objects.get(id=pk)
    
    if request.method == "POST":
        bomForm = BillModelForm(request.POST)
        request_data = request.POST.copy()
        mutable_data = QueryDict(mutable=True)
        mutable_data.update(request_data)
        if request_data['sales_price'] == "":
            mutable_data['sales_price'] = int(float(request_data['estimate_cost']) / 0.6)
        print(mutable_data)
        mutable_data.appendlist('project', project.pk)
        bomForm = BillModelForm(mutable_data)
        if bomForm.is_valid():
            bomForm.save()
            project_update(pk)
            print("created a new BOM")
            return HttpResponse("Saved bom")
    bomForm = BillModelForm(initial={'project': project})
    context = {"bomForm": bomForm, "project": project}
    return render(request, "budgetTool/partials/bomForm.html", context)

def create_service(request,pk):
    project=Project.objects.get(id=pk)
    rate_table=RateTableCost.objects.all()
    form=ServiceModelForm()
    if request.method =="POST":
        print(project.name)
        form=ServiceModelForm(request.POST,initial={'project': project})
        print(form.data)

        if form.is_valid():
            data=form.cleaned_data
            print(data)
            name=data['name']
            category=data['category']
            serviceType=data['type']
            hours_estimated=data['hours_estimated']
            hours_worked=data['hours_worked']
            # rate_list=data['rate_list']
            # rate_cost=data['rate_cost']
            travel_actual=data['travel_actual']
            isOnSite=data['isOnSite']
            notes=data['notes']
            
            for item in rate_table:
                if item.name == serviceType.name :
                    print(f'check match: name:{item.name}|--- type:{serviceType.name}|')
                    rate_cost=item.labor_cost
                    if isOnSite:
                        rate_list=item.labor_on_site
                    else:
                        rate_list=item.labor_in_house

            #calculated field value
            print(project.adjust_Service)
            if isOnSite:
                travel_estimate=hours_estimated*project.travel_weekly/40
            else:
                travel_estimate=0

            if travel_actual == None:
                travel_actual=0
            if hours_worked == None:
                hours_worked=0
            hours_adjusted=hours_estimated*(1+project.adjust_Service)
            sub_total_list=hours_estimated*rate_list+travel_estimate
            sub_total_adjusted_list=hours_adjusted*rate_list+travel_estimate
            sub_total_cost_est=hours_estimated*rate_cost+travel_estimate
            sub_total_adjusted_cost_est=hours_adjusted*rate_cost+travel_estimate

            cost_actual=hours_worked*rate_cost+travel_actual

            print(f'rate list: {rate_list}')
            Service.objects.create(   
                name=name,
                category=category,
                type=serviceType,
                hours_estimated=hours_estimated,
                hours_worked=hours_worked,
                rate_list=rate_list,
                rate_cost=rate_cost,
                travel_actual=travel_actual,
                isOnSite=isOnSite,
                project=project,
                hours_adjusted=hours_adjusted,
                travel_estimate=travel_estimate,
                sub_total_list=sub_total_list,
                sub_total_adjusted_list=sub_total_adjusted_list,
                sub_total_cost_est=sub_total_cost_est,
                sub_total_adjusted_cost_est=sub_total_adjusted_cost_est,
                cost_actual=cost_actual,
                notes=notes,
            )
            project_update(pk)
            return HttpResponse("Service Saved")
    context={
        "form":ServiceModelForm(initial={'project': project},),
        "project":project,
        "rate_cost":rate_table,
    }
    return render(request,"budgetTool/partials/serviceForm.html",context)

def service_edit(request, pk, fk):
    item = get_object_or_404(Service, id=pk, project_id=fk)
    project=Project.objects.get(id=fk)
    rate_table=RateTableCost.objects.all()

    if request.method == "POST":
        request_data = request.POST.copy()
        request_data['project'] = fk  # Add project to POST data
        form = ServiceModelForm(request_data, instance=item)

        if form.is_valid():
            data=form.cleaned_data
            item.name = data['name']
            item.category = data['category']
            item.type = data['type']
            item.hours_estimated = data['hours_estimated']
            item.hours_worked = data['hours_worked']
            item.rate_list = data['rate_list']
            item.rate_cost = data['rate_cost']
            item.travel_actual = data['travel_actual']
            item.isOnSite = data['isOnSite']
            item.notes = data['notes']

            for rate in rate_table:
                if item.type.name == rate.name:
                    item.rate_cost=rate.labor_cost
                    if item.isOnSite:
                        item.rate_list=rate.labor_on_site
                    else:
                        item.rate_list=rate.labor_in_house
                    
            #calculated field value
            print(project.adjust_Service)
            if item.isOnSite:
                item.travel_estimate=item.hours_estimated*project.travel_weekly/40
            else:
                item.travel_estimate=0

            if item.travel_actual == None:
                item.travel_actual=0
            if item.hours_worked == None:
                item.hours_worked=0

            item.hours_adjusted = item.hours_estimated * (1 + project.adjust_Service)
            item.sub_total_list = item.hours_estimated * item.rate_list + item.travel_estimate
            item.sub_total_adjusted_list = item.hours_adjusted * item.rate_list + item.travel_estimate
            item.sub_total_cost_est = item.hours_estimated * item.rate_cost + item.travel_estimate
            item.sub_total_adjusted_cost_est = item.hours_adjusted * item.rate_cost + item.travel_estimate
            item.cost_actual = item.hours_worked * item.rate_cost + item.travel_actual
            item.save()
            project_update(fk)
            return HttpResponse("Service updated successfully")

    else:
        form = ServiceModelForm(instance=item)

    context = {
        "item": item,
        "form": form,
    }
    return render(request, "budgetTool/partials/serviceFormEdit.html", context)

def service_copy(request, pk, fk):
    item = get_object_or_404(Service, id=pk, project_id=fk)
    project=Project.objects.get(id=fk)
    rate_table=RateTableCost.objects.all()

    if request.method == "POST":
        request_data = request.POST.copy()
        request_data['project'] = fk  # Add project to POST data
        form = ServiceModelForm(request_data)

        if form.is_valid():
            data=form.cleaned_data
            item.name = data['name']
            item.category = data['category']
            item.type = data['type']
            item.hours_estimated = data['hours_estimated']
            item.hours_worked = data['hours_worked']
            item.rate_list = data['rate_list']
            item.rate_cost = data['rate_cost']
            item.travel_actual = data['travel_actual']
            item.isOnSite = data['isOnSite']
            item.notes = data['notes']

            for rate in rate_table:
                if item.type.name == rate.name:
                    item.rate_cost=rate.labor_cost
                    if item.isOnSite:
                        item.rate_list=rate.labor_on_site
                    else:
                        item.rate_list=rate.labor_in_house
                    
            #calculated field value
            print(project.adjust_Service)
            if item.isOnSite:
                item.travel_estimate=item.hours_estimated*project.travel_weekly/40
            else:
                item.travel_estimate=0

            if item.travel_actual == None:
                item.travel_actual=0
            if item.hours_worked == None:
                item.hours_worked=0

            item.hours_adjusted = item.hours_estimated * (1 + project.adjust_Service)
            item.sub_total_list = item.hours_estimated * item.rate_list + item.travel_estimate
            item.sub_total_adjusted_list = item.hours_adjusted * item.rate_list + item.travel_estimate
            item.sub_total_cost_est = item.hours_estimated * item.rate_cost + item.travel_estimate
            item.sub_total_adjusted_cost_est = item.hours_adjusted * item.rate_cost + item.travel_estimate
            item.cost_actual = item.hours_worked * item.rate_cost + item.travel_actual
            item.pk=None
            item.save()
            project_update(fk)
            return HttpResponse("Service copy successfully")

    else:
        form = ServiceModelForm(instance=item)

    context = {
        "item": item,
        "form": form,
    }
    return render(request, "budgetTool/partials/serviceFormCopy.html", context)
 

def bom_edit(request, pk, fk):
    if request.user.groups.filter(name="Account Manager").exists():
        item = get_object_or_404(BillOfMaterials, id=pk, project_id=fk)
        if request.method == "POST":
            item.sales_price = request.POST.get('sales_price')
            item.quantity = request.POST.get('quantity')
            item.save(update_fields=["sales_price","quantity"])
            project_update(fk)
            return HttpResponse("BOM updated successfully")
        else:
            bomForm = BillModelForm(instance=item)
        context = {
            "item": item,
            "bomForm": bomForm,
        }
        return render(request, "budgetTool/accessControl/accountManager/bomFormEdit.html", context)



    item = get_object_or_404(BillOfMaterials, id=pk, project_id=fk)
    if request.method == "POST":
        request_data = request.POST.copy()
        request_data['project'] = fk  # Add project to POST data
        if request_data['sales_price'] == "":
            request_data['sales_price'] = int(float(request_data['estimate_cost']) / 0.6)
        bomForm = BillModelForm(request_data, instance=item)
        if bomForm.is_valid():
            bomForm.save()
            project_update(fk)
            return HttpResponse("BOM updated successfully")
    else:
        bomForm = BillModelForm(instance=item)
    context = {
        "item": item,
        "bomForm": bomForm,
    }
    return render(request, "budgetTool/partials/bomFormEdit.html", context)

def bom_copy(request, pk, fk):
    # if request.user.groups.filter(name="Account Manager").exists():
    #     item = get_object_or_404(BillOfMaterials, id=pk, project_id=fk)
    #     if request.method == "POST":
    #         item.sales_price = request.POST.get('sales_price')
    #         item.quantity = request.POST.get('quantity')
    #         item.save(update_fields=["sales_price","quantity"])
    #         project_update(fk)
    #         return HttpResponse("BOM updated successfully")
    #     else:
    #         bomForm = BillModelForm(instance=item)
    #     context = {
    #         "item": item,
    #         "bomForm": bomForm,
    #     }
    #     return render(request, "budgetTool/accessControl/accountManager/bomFormEdit.html", context)

    item = get_object_or_404(BillOfMaterials, id=pk, project_id=fk)
    if request.method == "POST":
        request_data = request.POST.copy()
        request_data['project'] = fk  # Add project to POST data
        if request_data['sales_price'] == "":
            request_data['sales_price'] = int(float(request_data['estimate_cost']) / 0.6)
        bomForm = BillModelForm(request_data)
        if bomForm.is_valid():
            new_bom =bomForm.save(commit=False)
            new_bom.pk=None
            new_bom.save()
            project_update(fk)
            return HttpResponse("BOM copy successfully")
    else:
        bomForm = BillModelForm(instance=item)
    context = {
        "item": item,
        "bomForm": bomForm,
    }
    return render(request, "budgetTool/partials/bomFormCopy.html", context)

def bom_delete(request,pk,fk):
    bom=BillOfMaterials.objects.get(id=pk)
    bom.delete()
    project_update(fk)
    return HttpResponse("BOM deleted successfully")

def service_delete(request,pk,fk):
    service=Service.objects.get(id=pk)
    service.delete()
    project_update(fk)
    return HttpResponse("Service deleted successfully")

def service_order(request,pk):
    form=OrderForm(request.POST)

    if form.is_valid():
        ordered_ids = form.cleaned_data["ordering"].split(',')
        user_profile = Profile.objects.get(user=request.user)
        user_profile.service_index = ordered_ids
        user_profile.save()
        print(ordered_ids)
        current_order = 1
        for lookup_id in ordered_ids:
            if lookup_id.isdigit() and lookup_id != "0":
                service = Service.objects.get(id=lookup_id)
                service.order = current_order
                service.save(update_fields=["order"])
                print(f'what is Service: {service.order}')
                current_order += 1
        return HttpResponse("Service order updated successfully")
    
def bom_order(request,pk):
    form=OrderForm(request.POST)

    if form.is_valid():
        ordered_ids = form.cleaned_data["ordering"].split(',')
        user_profile = Profile.objects.get(user=request.user)
        user_profile.bom_index = ordered_ids
        user_profile.save()
        current_order = 1
        for lookup_id in ordered_ids:
            if lookup_id.isdigit() and lookup_id != "0":
                bom = BillOfMaterials.objects.get(id=lookup_id)
                bom.order = current_order
                bom.save(update_fields=["order"])
                print(f'what is bom: {bom.order}')
                current_order += 1
        return HttpResponse("BOM order updated successfully")
    
def bom_category_edit(request,pk):
    category=BOMCategory.objects.get(id=pk)
    
    if request.method == 'POST':
        if request.POST.get('index'):
            category.index=request.POST.get('index','')
        if request.POST.get('name'):
            category.name=request.POST.get('name','')
        category.save()
        return HttpResponse("BOM category updated successfully")
    
    categoryForm=BomCategoryModelForm(instance=category)
    context = {"categoryForm": categoryForm, "category": category}
    if request.GET.get('index'):
        print(category.id)
        return render(request,'budgetTool/partials/category/editBomCategoryIndex.html',context)
    elif request.GET.get('name'):
        return render(request,'budgetTool/partials/category/editBomCategoryName.html',context)

def create_bom_category(request):
    form=BomCategoryModelForm()
    if request.method == "POST":
        form=BomCategoryModelForm(request.POST)
        if form.is_valid():
            print("valid")
            form.save()
            return HttpResponse("Category created successfully")
    context={"form":form}
    return render(request,'budgetTool/partials/category/createBomCategory.html',context)

def bom_category_delete(request,pk):
    category=BOMCategory.objects.get(id=pk)
    related_boms = BillOfMaterials.objects.filter(bom_category=category)
    if related_boms.exists():
        return HttpResponse("Cannot delete category. There are associated BillOfMaterials instances.")
    category.delete()
    return HttpResponse("Category deleted successfully")

def service_category_edit(request,pk):
    category=ServiceCategory.objects.get(id=pk)
    
    if request.method == 'POST':
        if request.POST.get('index'):
            category.index=request.POST.get('index','')
        if request.POST.get('name'):
            category.name=request.POST.get('name','')
        category.save()
        return HttpResponse("Service category updated successfully")
    
    categoryForm=ServiceCategoryModelForm(instance=category)
    context = {"categoryForm": categoryForm, "category": category}
    if request.GET.get('index'):
        return render(request,'budgetTool/partials/category/editServiceCategoryIndex.html',context)
    elif request.GET.get('name'):
        return render(request,'budgetTool/partials/category/editServiceCategoryName.html',context)
    
def create_service_category(request):
    form=ServiceCategoryModelForm()
    print("Confirm valid")
    if request.method == "POST":
        form=ServiceCategoryModelForm(request.POST)

        print(form.data)
        if form.is_valid():
            form.save()
            return HttpResponse("Category created successfully")
    context={"form":form}
    return render(request,'budgetTool/partials/category/createServiceCategory.html',context)

def apply_border_to_row(ws,row_index):
    border_style = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_index, max_row=row_index):
        for cell in col:
            cell.border = border_style

def download_excel(request,pk):
    project=Project.objects.get(id=pk)

    wb = Workbook()
    ws_summary = wb.create_sheet(title="Summary")
    
    ws_summary.append(["Summary"])

    data = [
        [" ","Cost Est","Cost Adjusted","List","List Adjusted","Quoted","Actual"],
        ["BOM","=BoM!E4","=BoM!E4","=BoM!G4","=BoM!H4","=BoM!D5","=BoM!F4"],
        ["Service","=Service!I4","=Service!J4","=Service!G4","=Service!H4","=Service!D5","=Service!K4"],
        ["Total","=C4+C5","=D4+D5","=E4+E5","=F4+F5","=G4+G5","=H4+H5"],
        ["Margin"," "," ","=1-C7/E7","=1-D7/F7","=1-D7/G7","=1-H7/G7"],
    ]

    for row in data:
        ws_summary.append(row)

    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=23)
        cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    for cell in ws_summary[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")        

    for i in range(1,7):
        apply_border_to_row(ws_summary, i)
    
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['A4'].font = Font(bold=True)
    ws_summary['A5'].font = Font(bold=True)
    ws_summary['A6'].font = Font(bold=True)
    
    ws_summary.insert_cols(1)
    ws_summary.insert_rows(1)
    ws_summary.merge_cells('B2:H2')

    for col_num in range(1, 8):
        ws_summary.column_dimensions[get_column_letter(col_num)].width = 15

    ws_summary['E7'].style = percentage_style
    ws_summary['F7'].style = percentage_style
    ws_summary['G7'].style = percentage_style
    ws_summary['H7'].style = percentage_style

    columns_to_format = ['C','D','E','F','G','H']
    for col in columns_to_format:
        for row in range(4, 7):
            cell = ws_summary[f'{col}{row}']
            cell.style = currency_style

    ws_summary.insert_rows(6)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    ws_summary.column_dimensions['A'].width = 2.5

    bom_page(wb,pk)
    service_page(wb,pk)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{project.name}-Project Budget v1.0.0.xlsx"'
    wb.save(response)

    return response

def bom_page(wb:Workbook, pk:int):
    project=Project.objects.get(id=pk)
    bom=BillOfMaterials.objects.filter(project_id=pk)
    bom_category=BOMCategory.objects.filter(project_BOM_category__in=bom).distinct()

    ws = wb.create_sheet(title="BOM")
    
    ws.append(["Totals"])
    summary = [
        ["BOM Variables","","Cost Estimate","Actual Cost","List","List Adjusted","Margin Est","Margin Adj","Margin Actual","Margin quoted"],
        ["Adjusted for complexity",f'{100*project.adjust_BOM} %',"=SUM(G10:G2000)","=SUM(J10:J2000)","=SUM(H10:H2000)","=G4*(1+D4)","=1-E4/G4","=1-E4/H4","=1-F4/D5","=1-E4/D5"],
        ["Quoted",project.quote_BOM,],
    ]

    for row in summary:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=23)
        cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    for cell in ws[2]:
        cell.alignment = Alignment(wrap_text=False)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")      

    for i in range(1,5):
        apply_border_to_row(ws, i)

    ws.insert_cols(1)
 
    data=[
        ["Bill of Materials"],
        ["#","Part/Item","Cost(Estimate)","List Price","Quantity",	"Total Cost(Estimate)",
         	"Total List","Supplier","Cost (Actual)","Responsible","Description","Notes"],
    ]

    for row in data:
        ws.append(row)

    for cell in ws[5]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=23)
        cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") 


    for category in bom_category:
        ws.append([category.name])
        ws[f'A{ws.max_row}'].font = Font(bold=True)
        bom_items = bom.filter(bom_category=category)
        for item in bom_items:
            ws.append([item.index,item.name,item.estimate_cost,item.sales_price,
                       item.quantity,item.estimate_cost*item.quantity,
                       item.sales_price*item.quantity,item.supplier,item.actual_cost,item.Responsible,item.description,item.notes]) 

    for i in range(5,ws.max_row+1):
        apply_border_to_row(ws, i)

    ws.insert_cols(1)
    ws.insert_rows(1)
    ws.insert_rows(6)

    ws.merge_cells('C2:L2')
    ws.merge_cells('B7:M7')
    ws.merge_cells('C3:D3')

    # Apply the style to the cell
    ws['D4'].style = percentage_style
    ws['I4'].style = percentage_style
    ws['J4'].style = percentage_style
    ws['K4'].style = percentage_style
    ws['L4'].style = percentage_style

    ws['E4'].style = currency_style
    ws['F4'].style = currency_style
    ws['G4'].style = currency_style
    ws['H4'].style = currency_style
    ws['D5'].style = currency_style

    columns_to_format = ['D','E','G','H','J']
    for col in columns_to_format:
        for row in range(10, ws.max_row+1):
            cell = ws[f'{col}{row}']
            cell.style = currency_style

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(cell.value) > max_length:
                    max_length = len(cell.value)
                    print(cell.value)
                    print(len(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length
    ws.column_dimensions['A'].width = 2.5

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return ws

def service_page(wb:Workbook, pk:int):
    project=Project.objects.get(id=pk)
    service=Service.objects.filter(project_id=pk)
    service_category=ServiceCategory.objects.filter(project_service_category__in=service).distinct()

    ws = wb.create_sheet(title="Service")
    ws.append(["Total System Integration Services"])
    summary = [
        ["Service Variables","","Hours","Risk Hours","Labor","Labor Adjusted","Cost Est","Cost Adjusted Est","Cost Act","Est Margin","Est Adjusted Margin","Actual Margin"],
        ["Adjusted for complexity",f'{100*project.adjust_Service} %',"=SUM(F10:F1000)","=SUM(G10:G1000)","=SUM(L10:L1000)","=SUM(M10:M1000)","=SUM(O10:O1000)","=SUM(P10:P1000)","=SUM(Q10:Q1000)","==1-I4/G4","==1-J4/H4","=1-K4/D5"],
        ["Quoted",project.quote_Service,],
        ["Travel Weekly",f'$ {project.travel_weekly}',],
    ]

    for row in summary:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=23)
        cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    for i in range(1,6):
        apply_border_to_row(ws, i)      

    ws.insert_cols(1)
 
    data=[
        ["Service"],
        ["#","Task","Type","On Site","Hours Estimated","Hours Adjusted","Hours Worked","Travel Estimate","Travel Actual","Rate List","Sub Total List","Sub Total Adjusted List","Rate Cost","Sub Total Cost Est","Sub Total Adjusted Cost Est","Cost Actual"],
    ]

    for row in data:
        ws.append(row)

    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=23)
        cell.fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    for cell in ws[7]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    isOnSite=""
    for current_category in service_category:
        ws.append([current_category.name])
        ws[f'A{ws.max_row}'].font = Font(bold=True)
        service_items = service.filter(category=current_category)
        for item in service_items:
            if item.isOnSite:
                isOnSite="X"
            else:
                isOnSite=""
            ws.append([item.index,item.name,item.type.name,isOnSite,item.hours_estimated,item.hours_adjusted,
                        item.hours_worked,item.travel_estimate,item.travel_actual,item.rate_list,item.sub_total_list,
                        item.sub_total_adjusted_list,item.rate_cost, item.sub_total_cost_est,item.sub_total_adjusted_cost_est,
                        item.cost_actual
                       ]) 

    for i in range(7,ws.max_row+1):
        apply_border_to_row(ws, i)

    ws.insert_cols(1)
    ws.insert_rows(1)
    ws.insert_rows(7)

    ws.merge_cells('C3:D3')
    ws.merge_cells('C2:N2')
    ws.merge_cells('D6:N6')
    ws.merge_cells('B8:Q8')
    # for col_num in range(1, 17):
    #     # ws.column_dimensions[get_column_letter(col_num)].width = 15
    #     ws.column_dimensions[get_column_letter(col_num)].bestFit = True
    

    ws['L4'].style = percentage_style
    ws['M4'].style = percentage_style
    ws['N4'].style = percentage_style

    ws['G4'].style = currency_style
    ws['H4'].style = currency_style
    ws['I4'].style = currency_style
    ws['J4'].style = currency_style
    ws['K4'].style = currency_style
    ws['D5'].style = currency_style

    columns_to_format = ['I', 'J', 'L', 'M', 'N', 'O', 'P', 'Q']
    for col in columns_to_format:
        for row in range(11, ws.max_row+1):
            cell = ws[f'{col}{row}']
            cell.style = currency_style

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(cell.value) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length
    
    ws.column_dimensions['A'].width = 2.5

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return ws
    
def create_rateTableCost(request):
    if request.method == "POST":
        form=RateCostModelForm(request.POST)
        request_data = request.POST.copy()
        mutable_data = QueryDict(mutable=True)
        mutable_data.update(request_data)

        base = int(request_data.get('base', 0))
        y1_value = (1.1 * base)
        y2_value = 1.1 * y1_value
        y3_value = 1.1 * y2_value
        y4_value = 1.3 * y3_value
        labor_cost = math.ceil(y4_value / 2080)
        mutable_data.appendlist('y1', int(y1_value))
        mutable_data.appendlist('y2', int(y2_value))
        mutable_data.appendlist('y3', int(y3_value))
        mutable_data.appendlist('y4', math.ceil(y4_value))
        mutable_data.appendlist('labor_cost', labor_cost)

        form=RateCostModelForm(mutable_data)
        print(form.data)
        if form.is_valid():
            form.save()
            print("created a new rate table cost item")
            return HttpResponse("Saved") 
    context = {"form": RateCostModelForm()}
    return render(request, "budgetTool/partials/rateCost/costFormCreate.html", context)

def rateCost_edit(request,pk):
    item = get_object_or_404(RateTableCost, id=pk)
    if request.method == "POST":
        request_data = request.POST.copy()
        base=int(request_data.get('base',0))
        y1_value=1.1*base
        y2_value=1.1*y1_value
        y3_value=1.1*y2_value
        y4_value=1.3*y3_value
        labor_cost=math.ceil(y4_value/2080)
        request_data['y1']=int(y1_value)
        request_data['y2']=int(y2_value)
        request_data['y3']=int(y3_value)
        request_data['y4']=math.ceil(y4_value)
        request_data['labor_cost']=labor_cost
        form = RateCostModelForm(request_data, instance=item)

        if form.is_valid():
            form.save()
            return HttpResponse("Rate Cost updated successfully")
    else:
        form = RateCostModelForm(instance=item)
        
    context = {
        "item": item,
        "form": form,
    }

    return render(request, "budgetTool/partials/rateCost/costFormEdit.html", context)

def rateCost_delete(request,pk):
    cost=RateTableCost.objects.get(id=pk)
    cost.delete()

@allowed_users(allowed_roles=['Admin','Account Manager','Engineering Lead'])
def price_sheet(request):
    vendors=Vendor.objects.all()
    user_profile=Profile.objects.get(user=request.user)
    table_order=user_profile.Product_Price_index
    if table_order == None:
        table_order = ""
    table_ids= [int(i) for i in table_order if i.isdigit()]
    table=Product_Price.objects.all()
    preserved_order = Case(
        *[When(pk=pk, then=Value(i)) for i, pk in enumerate(table_ids)],
        default=Value(len(table_ids)+1),
        output_field=IntegerField()
    )
    Product_Prices = table.annotate(custom_order=preserved_order).order_by('custom_order')
    
    item_length=0
    cost_length=0
    margin_length=0
    note_length=0

    for item in Product_Prices:
        if len(item.item) > item_length:
            item_length = len(item.item)
        if len(str(item.cost)) > cost_length:
            cost_length = len(str(item.cost))
        if len(str(item.margin)) > margin_length:
            margin_length = len(str(item.margin))
        if item.note and len(item.note) > note_length:
            note_length = len(item.note)

    print(item_length, cost_length, margin_length)

    if item_length < 10:
        item_length = 10
    if cost_length < 10:
        cost_length = 10
    if margin_length < 10:
        margin_length = 10
    if note_length < 15:
        note_length = 15
    elif note_length > 20:
        note_length = 20
    
    item_length *= 10
    cost_length *= 15
    margin_length *= 15
    note_length *= 35

    context={
        "vendors":vendors,
        "Product_Prices":Product_Prices,
        "item_length":item_length,
        "cost_length":cost_length,
        "margin_length":margin_length,
        "note_length":note_length,
    }

    return render(request,'budgetTool/price_sheet.html',context)

def create_product_price(request,fk):
    vendor=Vendor.objects.get(id=fk)

    if request.method == "POST":
        form=ProductPriceModelForm(request.POST)
        request_data = request.POST.copy()
        mutable_data = QueryDict(mutable=True)
        mutable_data.update(request_data)
        cost=float(request_data.get('cost',0))
        list=float(request_data.get('list',0))
        margin=(1-cost/list)*100
        rounded_margin = round(margin,2)
        mutable_data.appendlist('margin',rounded_margin)
        mutable_data.appendlist('index',100)
        mutable_data.appendlist('vendor',vendor.id)
        form=ProductPriceModelForm(mutable_data)
        print(form.data)
        if form.is_valid():
            form.save()
            return HttpResponse("Product Price created successfully")
    context={
        "form":ProductPriceModelForm(),
        "vendor":vendor,
    }
    return render(request,'budgetTool/partials/productPrice/priceFormCreate.html',context)

def product_price_edit(request,pk,fk):
    item = get_object_or_404(Product_Price, id=pk)
    vendor=Vendor.objects.get(id=fk)
    if request.method == "POST":
        request_data = request.POST.copy()
        cost=float(request_data.get('cost',0))
        list=float(request_data.get('list',0))
        margin=(1-cost/list)*100
        rounded_margin = round(margin,2)
        request_data['margin']=rounded_margin
        # mutable_data.appendlist('margin',rounded_margin)
        request_data.appendlist('vendor',vendor.id)
        request_data.appendlist('index',100)
        form=ProductPriceModelForm(request_data, instance=item) 
        print(form.data)

        if form.is_valid():
            print("create valid form")
            form.save()
            return HttpResponse("Product Price updated successfully")
    else:
        form = ProductPriceModelForm(instance=item)
        
    context = {
        "item": item,
        "form": form,
        "vendor":vendor,
    }
    return render(request, "budgetTool/partials/productPrice/priceFormEdit.html", context)

def product_price_copy(request,pk,fk):
    item = get_object_or_404(Product_Price, id=pk)
    vendor=Vendor.objects.get(id=fk)
    if request.method == "POST":
        request_data = request.POST.copy()
        cost=float(request_data.get('cost',0))
        list=float(request_data.get('list',0))
        margin=(1-cost/list)*100
        rounded_margin = round(margin,2)
        request_data['margin']=rounded_margin
        # mutable_data.appendlist('margin',rounded_margin)
        request_data.appendlist('vendor',vendor.id)
        request_data.appendlist('index',100)
        form=ProductPriceModelForm(request_data) 
        # print(form.data)

        if form.is_valid():
            print("valid form")
            new_item=form.save(commit=False)
            new_item.pk=None
            new_item.save()
            return HttpResponse("Product Price copy successfully")
    else:
        form = ProductPriceModelForm(instance=item)
        
    context = {
        "item": item,
        "form": form,
        "vendor":vendor,
    }
    return render(request, "budgetTool/partials/productPrice/priceFormCopy.html", context)

# def product_price_copy(request,pk):
#     item = get_object_or_404(Product_Price, id=pk)
#     item.pk=None
#     item.save()
#     return HttpResponse("Product Price copied successfully")


def product_price_delete(request,pk):
    price=Product_Price.objects.get(id=pk)
    price.delete()

def rate_cost_order(request):
    form=OrderForm(request.POST)
    if form.is_valid():
        ordered_ids = form.cleaned_data["ordering"].split(',')
        user_profile=Profile.objects.get(user=request.user)
        user_profile.RateTable_index=ordered_ids
        user_profile.save()
        return HttpResponse("order updated successfully")
    
def create_vendor(request):
    form=VendorModelForm()
    if request.method == "POST":
        form=VendorModelForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponse("Vendor created successfully")
    context={"form":form}
    return render(request,'budgetTool/partials/vendor/vendorForm.html',context)

def vendor_edit(request,pk):
    vendor=Vendor.objects.get(id=pk)
    if request.method == "POST":
        form=VendorModelForm(request.POST,instance=vendor)
        if form.is_valid():
            form.save()
            return HttpResponse("Vendor updated successfully")
    else:
        form=VendorModelForm(instance=vendor)
    context={
        "form":form,
        "vendor":vendor,
    }
    return render(request,'budgetTool/partials/vendor/vendorEdit.html',context)

def vendor_delete(request,pk):
    vendor=Vendor.objects.get(id=pk)
    vendor.delete()

def copy_project(request,pk):
    print("copy project")
    project=Project.objects.get(id=pk)
    project.pk=None
    project.name=project.name+" copy"
    project.isTemplate=False
    project.save()
    bom=BillOfMaterials.objects.filter(project_id=pk)
    service=Service.objects.filter(project_id=pk)
    for item in bom:
        item.pk=None
        item.project=project
        item.save()
    for item in service:
        item.pk=None
        item.project=project
        item.save()
    return redirect(f'/budgetTool/{project.id}')

def price_sheet_order(request):
    form=OrderForm(request.POST)

    if form.is_valid():
        ordered_ids = form.cleaned_data["ordering"].split(',')
        user_profile = Profile.objects.get(user=request.user)
        user_profile.Product_Price_index = ordered_ids
        user_profile.save()
        # print(ordered_ids)
        current_order = 1
        for lookup_id in ordered_ids:
            if lookup_id.isdigit() and lookup_id != "0":
                cost = Product_Price.objects.get(id=lookup_id)
                cost.order = current_order
                cost.save(update_fields=["order"])
                print(f'what is priceSheet index: {cost.order}')
                current_order += 1
        return HttpResponse("order updated successfully")

